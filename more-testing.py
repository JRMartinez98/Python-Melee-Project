import openpyxl as op

FrameData = {'-DR.MARIO-': 'Melee Project/Frame Data/Dr. Mario/',
                '-MARIO-' : 'Melee Project/Frame Data/Mario/',
                '-LUIGI-' : 'Melee Project/Frame Data/Luigi/',
                '-BOWSER-' : 'Melee Project/Frame Data/Bowser/',
                '-PEACH-' : 'Melee Project/Frame Data/Peach/',
                '-YOSHI-' : 'Melee Project/Frame Data/Yoshi/',
                '-DK-' : 'Melee Project/Frame Data/Donkey Kong/',
                '-FALCON-' : 'Melee Project/Frame Data/Captain Falcon/',
                '-GANON-' : 'Melee Project/Frame Data/Ganondorf/',
                '-FALCO-' : 'Melee Project/Frame Data/Falco/',
                '-FOX-' : 'Melee Project/Frame Data/Fox/',
                '-NESS-' : 'Melee Project/Frame Data/Ness/',
                '-ICE CLIMBERS-' : 'Melee Project/Frame Data/Ice Climbers/',
                '-KIRBY-' : 'Melee Project/Frame Data/Kirby/',
                '-SAMUS-' : 'Melee Project/Frame Data/Samus/',
                '-ZELDA-' : 'Melee Project/Frame Data/Zelda/',
                '-LINK-' : 'Melee Project/Frame Data/Link/',
                '-YLINK-' : 'Melee Project/Frame Data/Young Link/',
                '-PICHU-' : 'Melee Project/Frame Data/Pichu/',
                '-PIKACHU-' : 'Melee Project/Frame Data/Pikachu/',
                '-PUFF-' : 'Melee Project/Frame Data/Jigglypuff/',
                '-MEWTWO-' : 'Melee Project/Frame Data/Mewtwo/',
                '-G&W-' : 'Melee Project/Frame Data/Mr. Game & Watch/',
                '-MARTH-' : 'Melee Project/Frame Data/Marth/',
                '-ROY-' : 'Melee Project/Frame Data/Roy/',
                '-SHEIK-' : 'Melee Project/Frame Data/Sheik/'
}

ExcelCol = {'-DR.MARIO-': 'A',
                '-MARIO-' : 'B',
                '-LUIGI-' : 'C',
                '-BOWSER-' : 'D',
                '-PEACH-' : 'E',
                '-YOSHI-' : 'F',
                '-DK-' : 'G',
                '-FALCON-' : 'H',
                '-GANON-' : 'I',
                '-FALCO-' : 'J',
                '-FOX-' : 'K',
                '-NESS-' : 'L',
                '-ICE CLIMBERS-' : 'M',
                '-KIRBY-' : 'N',
                '-SAMUS-' : 'O',
                '-ZELDA-' : 'P',
                '-LINK-' : 'Q',
                '-YLINK-' : 'R',
                '-PICHU-' : 'S',
                '-PIKACHU-' : 'T',
                '-PUFF-' : 'U',
                '-MEWTWO-' : 'V',
                '-G&W-' : 'W',
                '-MARTH-' : 'X',   
                '-ROY-' : 'Y',
                '-SHEIK-' : 'Z',      
}

path = "C:\\Users\\Jose\\Desktop\\Melee-Project.xlsx"

wb_obj = op.load_workbook(path)
sheet_obj = wb_obj.active


cell_obj = sheet_obj.cell(row = 2, column = 1)

#print(cell_obj.value)
print(sheet_obj.max_row)


files = []
paths = []
character = ''
i = 0
for cell in sheet_obj['Q']:
    
    if i > 0 and cell.value != None:
        files.append(cell.value)
    
    elif i == 0:
        character = cell.value
    
    i = i + 1

for img in files:
    paths.append(FrameData[character]+img)

    
#for i in files:
#    print(i)



for i in paths:
    print(i)


print(FrameData[character])
